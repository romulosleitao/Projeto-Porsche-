"""
Porsche Sales Sanitization Agent
================================

Lê um arquivo .xlsx de vendas Porsche (cru) e aplica as regras definidas em
schema.md para gerar um novo arquivo tratado, com cada coluna sanitizada
inserida logo após a sua coluna fonte.

Uso:
    python sanitize_porsche.py [INPUT_XLSX] [OUTPUT_XLSX]

Sem argumentos, usa os defaults definidos em INPUT_PATH e OUTPUT_PATH abaixo.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------
INPUT_PATH = Path("porsche_database.xlsx")
OUTPUT_PATH = Path("porsche_database_sanitized.xlsx")
INVALID = "INVALID"


# ---------------------------------------------------------------------------
# Date sanitization
# ---------------------------------------------------------------------------
MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def _safe_date(y: int, m: int, d: int) -> str:
    try:
        return dt.date(y, m, d).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return INVALID


def sanitize_date(value: Any) -> str:
    if value is None or value == "":
        return INVALID

    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")

    s = str(value).strip()
    if not s:
        return INVALID

    # YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", s)
    if m:
        return _safe_date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    # MM/DD/YYYY or MM-DD-YYYY
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)
    if m:
        return _safe_date(int(m.group(3)), int(m.group(1)), int(m.group(2)))

    # MM/DD/YY or MM-DD-YY
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{2})$", s)
    if m:
        yy = int(m.group(3))
        year = 2000 + yy if yy < 70 else 1900 + yy
        return _safe_date(year, int(m.group(1)), int(m.group(2)))

    # Month DDth, YYYY  /  Mon DDth YYYY  /  Month DD, YYYY
    m = re.match(
        r"^([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?[,\s]+(\d{4})$",
        s,
    )
    if m:
        mon_key = m.group(1).lower().rstrip(".")
        if mon_key in MONTHS:
            return _safe_date(int(m.group(3)), MONTHS[mon_key], int(m.group(2)))

    return INVALID


# ---------------------------------------------------------------------------
# Porsche model sanitization
# ---------------------------------------------------------------------------
CANONICAL_MODELS = [
    "911 Carrera", "911 Carrera S", "911 Carrera GTS",
    "911 Turbo", "911 Turbo S",
    "911 GT3", "911 GT3 RS",
    "911 Dakar",
    "911 Targa 4", "911 Targa 4S",
    "718 Cayman", "718 Cayman S", "718 Cayman GT4 RS",
    "718 Boxster", "718 Boxster GTS",
    "718 Spyder RS",
    "Cayenne", "Cayenne S", "Cayenne Coupe", "Cayenne E-Hybrid",
    "Cayenne Turbo", "Cayenne Turbo GT",
    "Macan", "Macan S", "Macan T", "Macan GTS", "Macan Electric",
    "Panamera", "Panamera 4", "Panamera 4S",
    "Panamera Turbo", "Panamera Turbo S", "Panamera 4 E-Hybrid",
    "Taycan", "Taycan 4S", "Taycan GTS",
    "Taycan Turbo", "Taycan Turbo S", "Taycan Cross Turismo",
]

# Lookup case-insensitive por chave normalizada
_MODEL_LOOKUP = {m.lower(): m for m in CANONICAL_MODELS}


def _smart_title(s: str) -> str:
    """Title case preservando palavras comuns como E-Hybrid, GTS, RS, etc."""
    keep_upper = {"GT", "GTS", "RS", "GT3", "GT4", "S", "T", "4", "4S", "USA"}
    parts = s.split()
    out = []
    for p in parts:
        token = p.upper()
        if token in keep_upper:
            out.append(token)
        elif "-" in p:
            out.append("-".join(seg.capitalize() for seg in p.split("-")))
        else:
            out.append(p.capitalize())
    return " ".join(out)


def sanitize_model(value: Any) -> str:
    if value is None:
        return INVALID
    s = " ".join(str(value).split()).strip()
    if not s:
        return INVALID
    key = s.lower()
    if key in _MODEL_LOOKUP:
        return _MODEL_LOOKUP[key]
    # Modelos desconhecidos ficam title-case
    return _smart_title(s)


# ---------------------------------------------------------------------------
# Model year sanitization
# ---------------------------------------------------------------------------
WORDS_TO_NUM = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4,
    "five": 5, "six": 6, "seven": 7, "eight": 8, "nine": 9,
    "ten": 10, "eleven": 11, "twelve": 12, "thirteen": 13,
    "fourteen": 14, "fifteen": 15, "sixteen": 16, "seventeen": 17,
    "eighteen": 18, "nineteen": 19, "twenty": 20, "thirty": 30,
    "forty": 40, "fifty": 50, "sixty": 60, "seventy": 70,
    "eighty": 80, "ninety": 90,
    "hundred": 100, "thousand": 1000,
}


def _words_to_int(text: str) -> int | None:
    """Converte 'twenty twenty four', 'two thousand twenty two', 'eighty two thousand', etc."""
    tokens = re.findall(r"[a-z]+", text.lower())
    if not tokens:
        return None
    if any(t not in WORDS_TO_NUM for t in tokens):
        return None

    # Trata caso "twenty twenty four" (estilo ano: pares de dezenas)
    # quando todas tokens são <= 90 e nenhuma é hundred/thousand
    if all(WORDS_TO_NUM[t] < 100 for t in tokens):
        # tenta ler como dois pares: "twenty twenty four" -> 20*100 + 24
        nums = [WORDS_TO_NUM[t] for t in tokens]
        # combinar pares: ex [20, 20, 4] -> dois números: 20 e 24
        groups: list[int] = []
        i = 0
        while i < len(nums):
            if i + 1 < len(nums) and nums[i] >= 20 and nums[i] % 10 == 0 and nums[i + 1] < 10:
                groups.append(nums[i] + nums[i + 1])
                i += 2
            else:
                groups.append(nums[i])
                i += 1
        if len(groups) == 2:
            return groups[0] * 100 + groups[1]
        if len(groups) == 1:
            return groups[0]

    # Estilo "two thousand twenty two" / "eighty two thousand"
    total = 0
    current = 0
    for t in tokens:
        v = WORDS_TO_NUM[t]
        if v == 100:
            current = max(1, current) * 100
        elif v == 1000:
            current = max(1, current) * 1000
            total += current
            current = 0
        else:
            current += v
    return total + current


def sanitize_year(value: Any) -> str:
    if value is None:
        return INVALID

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        y = int(value)
        return str(y) if 1990 <= y <= 2035 else INVALID

    s = str(value).strip()
    if not s:
        return INVALID

    # 4 dígitos
    m = re.fullmatch(r"\d{4}", s)
    if m:
        y = int(s)
        return str(y) if 1990 <= y <= 2035 else INVALID

    # Formatos como 20-23, 20 24, 20.24
    m = re.fullmatch(r"(\d{2})\s*[-\s.]\s*(\d{2})", s)
    if m:
        y = int(m.group(1) + m.group(2))
        return str(y) if 1990 <= y <= 2035 else INVALID

    # Texto "twenty twenty four"
    if re.search(r"[A-Za-z]", s):
        n = _words_to_int(s)
        if n is not None and 1990 <= n <= 2035:
            return str(n)

    return INVALID


# ---------------------------------------------------------------------------
# Sales price sanitization
# ---------------------------------------------------------------------------
PRICE_WORD_MULTIPLIERS = {
    "thousand": 1_000,
    "million": 1_000_000,
}


def _parse_textual_price(s: str) -> float | None:
    """Parses 'eighty two thousand', 'two hundred thousand', etc."""
    tokens = re.findall(r"[a-z]+", s.lower())
    if not tokens:
        return None
    valid = set(WORDS_TO_NUM) | {"million"}
    if any(t not in valid for t in tokens):
        return None

    total = 0
    current = 0
    for t in tokens:
        if t == "million":
            current = max(1, current) * 1_000_000
            total += current
            current = 0
        elif t == "hundred":
            current = max(1, current) * 100
        elif t == "thousand":
            current = max(1, current) * 1_000
            total += current
            current = 0
        else:
            current += WORDS_TO_NUM[t]
    return float(total + current)


def _parse_numeric_price(num_str: str) -> float | None:
    """Lida com formatos US e europeus."""
    s = num_str.strip()
    if not s:
        return None

    has_dot = "." in s
    has_comma = "," in s

    if has_dot and has_comma:
        # Last separator wins as decimal
        if s.rfind(",") > s.rfind("."):
            # europeu: . é milhar, , é decimal
            s = s.replace(".", "").replace(",", ".")
        else:
            # US: , é milhar, . é decimal
            s = s.replace(",", "")
    elif has_comma:
        # Apenas vírgula. Se houver mais de uma OU os dígitos pós-vírgula não forem 2,
        # tratamos como milhar.
        last = s.rsplit(",", 1)
        digits_after = re.sub(r"\D", "", last[1])
        if s.count(",") > 1 or len(digits_after) != 2:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif has_dot:
        # Apenas ponto. Se múltiplos pontos OU dígitos após o último não forem 2,
        # tratamos como milhar (estilo europeu).
        last = s.rsplit(".", 1)
        digits_after = re.sub(r"\D", "", last[1])
        if s.count(".") > 1 or len(digits_after) != 2:
            s = s.replace(".", "")

    try:
        return float(s)
    except ValueError:
        return None


def sanitize_price(value: Any) -> str:
    if value is None or value == "":
        return INVALID

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"{float(value):.2f}"

    s = str(value).strip()
    if not s:
        return INVALID

    # Remover símbolos comuns e palavras-chave (USD, dollars)
    cleaned = s.replace("$", " ")
    cleaned = re.sub(r"\b(usd|dollars?)\b", " ", cleaned, flags=re.IGNORECASE)
    cleaned = cleaned.strip()

    # Preço textual: contém apenas letras e espaços
    if re.fullmatch(r"[A-Za-z\s]+", cleaned):
        n = _parse_textual_price(cleaned)
        return f"{n:.2f}" if n is not None else INVALID

    # Multiplicador "k" (mil)
    multiplier = 1.0
    m = re.search(r"([0-9.,]+)\s*k\b", cleaned, flags=re.IGNORECASE)
    if m:
        num_part = m.group(1)
        multiplier = 1000.0
    else:
        # Pega o token numérico
        m = re.search(r"[0-9.,]+", cleaned)
        if not m:
            return INVALID
        num_part = m.group(0)

    n = _parse_numeric_price(num_part)
    if n is None:
        return INVALID
    return f"{n * multiplier:.2f}"


# ---------------------------------------------------------------------------
# Vehicle mileage sanitization
# ---------------------------------------------------------------------------
def sanitize_mileage(value: Any) -> str:
    if value is None:
        return INVALID

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Pode vir como float (ex.: 9.5). Tratamos como inteiro de milhas.
        return str(int(round(float(value))))

    s = str(value).strip()
    if not s:
        return INVALID

    low = s.lower()
    if low in {"new", "new car", "zero miles", "zero", "0 mi", "0 miles", "0mi"}:
        return "0"

    # KM (precisa converter)
    is_km = bool(re.search(r"\bkm\b", low) or low.startswith("km"))

    # Texto puro (sem dígitos): "twelve thousand miles", "fifteen thousand miles"
    digits_present = bool(re.search(r"\d", s))
    if not digits_present:
        # Remover unidade
        textual = re.sub(r"\b(mi|miles|mile|km|kilometers|kilometres)\b", " ", low)
        textual = textual.strip()
        n = _parse_textual_price(textual)  # mesma lógica de palavras
        if n is None:
            return INVALID
        miles = n * (0.621371 if is_km else 1.0)
        return str(int(round(miles)))

    # Capturar parte numérica
    m = re.search(r"[0-9][0-9.,]*", s)
    if not m:
        return INVALID
    num_part = m.group(0)
    n = _parse_numeric_price(num_part)
    if n is None:
        return INVALID
    if is_km:
        n = n * 0.621371
    return str(int(round(n)))


# ---------------------------------------------------------------------------
# Payment method sanitization
# ---------------------------------------------------------------------------
PAYMENT_CANONICAL = [
    "Credit Card", "Debit Card", "Bank Transfer", "Wire Transfer",
    "Financing", "Lease", "Cash", "ACH Payment", "Crypto Payment",
]


def _normalize_text_key(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[\W_]+", " ", s.lower()).strip()
    s = re.sub(r"\s+", " ", s)
    return s


PAYMENT_LOOKUP = {
    # Credit Card
    "credit card": "Credit Card",
    "creditcard": "Credit Card",
    "credit": "Credit Card",
    "credit card payment": "Credit Card",
    # Debit Card
    "debit card": "Debit Card",
    "debitcard": "Debit Card",
    # Bank Transfer
    "bank transfer": "Bank Transfer",
    "banktransfer": "Bank Transfer",
    "bank": "Bank Transfer",
    # Wire Transfer
    "wire transfer": "Wire Transfer",
    "wiretransfer": "Wire Transfer",
    "wire": "Wire Transfer",
    "bank wire": "Wire Transfer",
    "wire bank": "Wire Transfer",
    # Financing
    "financing": "Financing",
    "finance": "Financing",
    "financing plan": "Financing",
    # Lease
    "lease": "Lease",
    "leasing": "Lease",
    "lease plan": "Lease",
    # Cash
    "cash": "Cash",
    "cash payment": "Cash",
    # ACH
    "ach": "ACH Payment",
    "ach payment": "ACH Payment",
    # Crypto
    "crypto": "Crypto Payment",
    "crypto payment": "Crypto Payment",
    "cryptocurrency": "Crypto Payment",
}


def sanitize_payment(value: Any) -> str:
    if value is None:
        return INVALID
    s = str(value).strip()
    if not s:
        return INVALID
    key = _normalize_text_key(s)
    if key in PAYMENT_LOOKUP:
        return PAYMENT_LOOKUP[key]
    # Desconhecidos: title case
    return " ".join(w.capitalize() for w in key.split())


# ---------------------------------------------------------------------------
# City sanitization
# ---------------------------------------------------------------------------
def sanitize_city(value: Any) -> str:
    if value is None:
        return INVALID
    s = " ".join(str(value).split()).strip()
    if not s:
        return INVALID

    parts = s.split(" ")
    out: list[str] = []
    for p in parts:
        if p.lower() == "st." or p.lower() == "st":
            out.append("St.")
        elif p.lower() == "mt." or p.lower() == "mt":
            out.append("Mt.")
        elif "-" in p:
            out.append("-".join(seg.capitalize() for seg in p.split("-")))
        else:
            out.append(p.capitalize())
    return " ".join(out)


# ---------------------------------------------------------------------------
# State sanitization
# ---------------------------------------------------------------------------
US_STATES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN",
    "mississippi": "MS", "missouri": "MO", "montana": "MT", "nebraska": "NE",
    "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR",
    "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC",
}
US_STATE_CODES = set(US_STATES.values())


def sanitize_state(value: Any) -> str:
    if value is None:
        return INVALID
    s = str(value).strip()
    if not s:
        return INVALID
    upper = s.upper().strip()
    if upper in US_STATE_CODES:
        return upper
    key = " ".join(s.lower().split())
    if key in US_STATES:
        return US_STATES[key]
    return INVALID


# ---------------------------------------------------------------------------
# Delivery status sanitization
# ---------------------------------------------------------------------------
DELIVERY_LOOKUP = {
    "delivered": "Delivered",
    "deliverd": "Delivered",  # typo comum
    "pending": "Pending",
    "in transit": "In Transit",
    "intransit": "In Transit",
    "cancelled": "Cancelled",
    "canceled": "Cancelled",
    "awaiting delivery": "Awaiting Delivery",
    "awaiting pickup": "Awaiting Pickup",
    "pending approval": "Pending Approval",
    "pending review": "Pending Review",
    "shipped": "Shipped",
    "awaiting review": "Awaiting Review",
}


def sanitize_delivery(value: Any) -> str:
    if value is None:
        return INVALID
    s = str(value).strip()
    if not s:
        return INVALID
    key = _normalize_text_key(s)
    if key in DELIVERY_LOOKUP:
        return DELIVERY_LOOKUP[key]
    return " ".join(w.capitalize() for w in key.split()) or INVALID


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------
COLUMN_PIPELINE = [
    # (input_col, sanitized_col_name or None, sanitizer or None)
    ("sale_id", None, None),
    ("sale_date", "SaleDateSanitized", sanitize_date),
    ("customer_name", None, None),
    ("porsche_model", "PorscheModelSanitized", sanitize_model),
    ("model_year", "ModelYearSanitized", sanitize_year),
    ("sale_price", "SalesPriceSanitized", sanitize_price),
    ("vehicle_mileage", "VehicleMileageSanitized", sanitize_mileage),
    ("payment_method", "PayMethodSanitized", sanitize_payment),
    ("city", "CitySanitized", sanitize_city),
    ("state", "StateSanitized", sanitize_state),
    ("salesperson", None, None),
    ("delivery_status", "DeliveryStatusSanitized", sanitize_delivery),
]


def process_workbook(input_path: Path, output_path: Path) -> dict[str, int]:
    wb_in = load_workbook(input_path, data_only=True)
    ws_in = wb_in.active

    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty workbook")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    name_to_idx = {h: i for i, h in enumerate(header)}

    # Validar presença das colunas requeridas
    missing = [c for c, _, _ in COLUMN_PIPELINE if c not in name_to_idx]
    if missing:
        raise ValueError(f"Missing required columns in input: {missing}")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sanitized"

    # Cabeçalho final
    out_header: list[str] = []
    for src, san_name, _ in COLUMN_PIPELINE:
        out_header.append(src)
        if san_name:
            out_header.append(san_name)
    ws_out.append(out_header)

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    sanitized_fill = PatternFill("solid", fgColor="065F46")
    for col_idx, name in enumerate(out_header, start=1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = sanitized_fill if name.endswith("Sanitized") else header_fill

    # Estatísticas de invalidos
    invalid_counts: dict[str, int] = {}

    for r in rows[1:]:
        out_row: list[Any] = []
        for src, san_name, fn in COLUMN_PIPELINE:
            raw = r[name_to_idx[src]]
            out_row.append(raw)
            if san_name and fn is not None:
                clean = fn(raw)
                if clean == INVALID:
                    invalid_counts[san_name] = invalid_counts.get(san_name, 0) + 1
                out_row.append(clean)
        ws_out.append(out_row)

    # Largura das colunas
    for col_idx in range(1, len(out_header) + 1):
        letter = get_column_letter(col_idx)
        ws_out.column_dimensions[letter].width = max(14, len(out_header[col_idx - 1]) + 2)
    ws_out.freeze_panes = "A2"

    wb_out.save(output_path)
    return invalid_counts


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", nargs="?", default=str(INPUT_PATH))
    parser.add_argument("output", nargs="?", default=str(OUTPUT_PATH))
    args = parser.parse_args(argv)

    in_path = Path(args.input)
    out_path = Path(args.output)
    if not in_path.exists():
        print(f"ERROR: input file not found: {in_path}", file=sys.stderr)
        return 1

    stats = process_workbook(in_path, out_path)
    print(f"OK -> {out_path}")
    if stats:
        print("INVALID counts per sanitized column:")
        for k, v in sorted(stats.items()):
            print(f"  {k}: {v}")
    else:
        print("No INVALID values produced.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
